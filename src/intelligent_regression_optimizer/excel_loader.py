"""Excel import adapter for SuiteCompass.

Reads a test inventory spreadsheet and converts it to a list of test dicts
matching the SuiteCompass input schema.
"""
from __future__ import annotations

import pathlib
import re
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

VALID_LAYERS = {"e2e", "integration", "unit", "security", "performance"}

# Canonical column names and their fuzzy-match aliases (normalised to lowercase,
# stripped of spaces/underscores/parens/hyphens).
_CANONICAL_MAP: dict[str, str] = {
    # target field -> canonical display header
    "id":                     "ID",
    "name":                   "Name",
    "layer":                  "Layer",
    "coverage_areas":         "Coverage Areas",
    "execution_time_secs":    "Execution Time (secs)",
    "flakiness_rate":         "Flakiness Rate",
    "failure_count_last_30d": "Failure Count (30d)",
    "automated":              "Automated",
    "tags":                   "Tags",
}

# All accepted aliases per field (normalised keys)
_ALIASES: dict[str, list[str]] = {
    "id":                     ["id"],
    "name":                   ["name"],
    "layer":                  ["layer", "testlayer", "type"],
    "coverage_areas":         ["coverageareas", "coveragearea", "areas", "coverage"],
    "execution_time_secs":    ["executiontimesecs", "executiontime", "exectime",
                               "durationinseconds", "duration", "timesecs"],
    "flakiness_rate":         ["flakinessrate", "flakiness", "flakyrate"],
    "failure_count_last_30d": ["failurecount30d", "failurecountlast30d", "failurecount",
                               "failures30d", "failures"],
    "automated":              ["automated", "isautomated", "auto"],
    "tags":                   ["tags", "tag", "labels", "label"],
}

REQUIRED_FIELDS = {"id", "name", "layer", "coverage_areas", "execution_time_secs", "flakiness_rate"}


# ---------------------------------------------------------------------------
# Errors
# ---------------------------------------------------------------------------

class ExcelLoaderError(ValueError):
    """Raised when the Excel file cannot be loaded or fails validation."""


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _normalise(text: str) -> str:
    """Strip and lowercase, remove spaces/underscores/hyphens/parens."""
    return re.sub(r"[\s_()\-]+", "", str(text)).lower()


def _match_header(raw_header: str) -> str | None:
    """Return the field name for a given raw header, or None if unrecognised."""
    normalised = _normalise(raw_header)
    # Strict: direct match against canonical display name
    for field, canonical in _CANONICAL_MAP.items():
        if normalised == _normalise(canonical):
            return field
    # Fuzzy: alias lookup
    for field, aliases in _ALIASES.items():
        if normalised in aliases:
            return field
    return None


def _find_header_row(ws, max_scan: int = 5) -> int | None:
    """Return the 1-based row index of the header row (containing >= 3 known columns).

    Scans up to *max_scan* rows from the top.
    """
    for row_idx in range(1, max_scan + 1):
        row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        matched = sum(1 for v in row_values if v and _match_header(str(v)) is not None)
        if matched >= 3:
            return row_idx
    return None


def _detect_merged_cells(ws) -> list[str]:
    """Return descriptions of any merged cell ranges that overlap data rows."""
    return [str(r) for r in ws.merged_cells.ranges]


def _parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    s = str(value).strip().lower()
    if s in {"true", "yes", "1", "y"}:
        return True
    if s in {"false", "no", "0", "n"}:
        return False
    raise ValueError(f"Cannot parse as boolean: {value!r}")


def _parse_list(value: Any) -> list[str]:
    if value is None or str(value).strip() == "":
        return []
    return [item.strip() for item in str(value).split(",") if item.strip()]


def _select_sheet(wb: openpyxl.Workbook, sheet: str | None) -> openpyxl.worksheet.worksheet.Worksheet:
    """Return the target worksheet according to selection rules."""
    if sheet is not None:
        if sheet in wb.sheetnames:
            return wb[sheet]
        raise ExcelLoaderError(
            f"Sheet {sheet!r} not found. Available sheets: {wb.sheetnames}"
        )
    # Auto-detect: prefer sheet named "Tests" (case-insensitive)
    for name in wb.sheetnames:
        if name.strip().lower() == "tests":
            return wb[name]
    # Fall back to first sheet
    return wb[wb.sheetnames[0]]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def load_excel(path: str, sheet: str | None = None) -> list[dict[str, Any]]:
    """Load a test inventory from an Excel file.

    Args:
        path:  Path to the ``.xlsx`` file.
        sheet: Sheet name to read. If omitted, auto-selects a sheet named
               "Tests" (case-insensitive), falling back to the first sheet.

    Returns:
        A list of test dicts matching the SuiteCompass ``test_suite`` schema.

    Raises:
        :class:`ExcelLoaderError` for any load or validation failure.
    """
    p = pathlib.Path(path)

    if not p.exists():
        raise ExcelLoaderError(f"File not found: {path!r}")

    if p.suffix.lower() != ".xlsx":
        raise ExcelLoaderError(
            f"Unsupported file format: {p.suffix!r}. Use .xlsx format."
        )

    try:
        wb = openpyxl.load_workbook(str(p), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ExcelLoaderError(f"Cannot open file {path!r}: {exc}") from exc

    ws = _select_sheet(wb, sheet)

    # Detect merged cells
    merged = _detect_merged_cells(ws)
    if merged:
        raise ExcelLoaderError(
            f"Merged cells detected in sheet {ws.title!r}: {merged[0]}. "
            f"Unmerge all cells before importing."
        )

    # Locate header row
    header_row_idx = _find_header_row(ws)
    if header_row_idx is None:
        raise ExcelLoaderError(
            f"Could not find a header row in sheet {ws.title!r}. "
            f"Ensure headers include at least 3 of: {list(_CANONICAL_MAP.values())}"
        )

    # Build column → field mapping
    col_to_field: dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=header_row_idx, column=col).value
        if cell_val is None:
            continue
        field = _match_header(str(cell_val))
        if field:
            col_to_field[col] = field

    # Check all required fields are present
    found_fields = set(col_to_field.values())
    missing = REQUIRED_FIELDS - found_fields
    if missing:
        # Provide canonical names for missing fields
        missing_names = [_CANONICAL_MAP[f] for f in sorted(missing)]
        raise ExcelLoaderError(
            f"Missing required column(s): {missing_names}. "
            f"Check column headers in sheet {ws.title!r}."
        )

    # Parse data rows
    tests: list[dict[str, Any]] = []
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        # Stop at first fully-empty row
        row_values = {
            col_to_field[col]: ws.cell(row=row_idx, column=col).value
            for col in col_to_field
        }
        if all(v is None or str(v).strip() == "" for v in row_values.values()):
            break

        test = _parse_row(row_values, row_idx)
        tests.append(test)

    if not tests:
        raise ExcelLoaderError(
            f"Sheet {ws.title!r} has headers but no data rows."
        )

    return tests


def _parse_row(raw: dict[str, Any], row_idx: int) -> dict[str, Any]:
    """Parse and validate a single data row. row_idx is 1-based (for error messages)."""
    # --- id ---
    id_val = str(raw.get("id", "") or "").strip()
    if not id_val:
        raise ExcelLoaderError(f"Row {row_idx}, column 'ID': value is empty. ID is required.")

    # --- name ---
    name_val = str(raw.get("name", "") or "").strip()
    if not name_val:
        raise ExcelLoaderError(f"Row {row_idx}, column 'Name': value is empty. Name is required.")

    # --- layer ---
    layer_val = str(raw.get("layer", "") or "").strip().lower()
    if layer_val not in VALID_LAYERS:
        raise ExcelLoaderError(
            f"Row {row_idx}, column 'Layer': value {str(raw.get('layer'))!r} is not valid. "
            f"Must be one of: {sorted(VALID_LAYERS)}"
        )

    # --- coverage_areas ---
    coverage_areas = _parse_list(raw.get("coverage_areas"))

    # --- execution_time_secs ---
    try:
        exec_time = float(raw.get("execution_time_secs") or 0)
    except (TypeError, ValueError):
        raise ExcelLoaderError(
            f"Row {row_idx}, column 'Execution Time (secs)': "
            f"value {raw.get('execution_time_secs')!r} must be a number."
        )
    if exec_time < 0:
        raise ExcelLoaderError(
            f"Row {row_idx}, column 'Execution Time (secs)': "
            f"value {exec_time} must be non-negative."
        )

    # --- flakiness_rate ---
    try:
        flakiness = float(raw.get("flakiness_rate") or 0)
    except (TypeError, ValueError):
        raise ExcelLoaderError(
            f"Row {row_idx}, column 'Flakiness Rate': "
            f"value {raw.get('flakiness_rate')!r} must be a number."
        )
    if not 0.0 <= flakiness <= 1.0:
        raise ExcelLoaderError(
            f"Row {row_idx}, column 'Flakiness Rate': "
            f"value {flakiness} must be between 0.0 and 1.0."
        )

    # --- failure_count_last_30d (optional) ---
    try:
        failure_count = int(raw.get("failure_count_last_30d") or 0)
    except (TypeError, ValueError):
        failure_count = 0

    # --- automated (optional) ---
    try:
        automated = _parse_bool(raw.get("automated", True))
    except ValueError:
        automated = True

    # --- tags (optional) ---
    tags = _parse_list(raw.get("tags"))

    return {
        "id": id_val,
        "name": name_val,
        "layer": layer_val,
        "coverage_areas": coverage_areas,
        "execution_time_secs": exec_time,
        "flakiness_rate": flakiness,
        "failure_count_last_30d": failure_count,
        "automated": automated,
        "tags": tags,
    }
