"""Tests for `iro import-tests` CLI subcommand — written RED before implementation. (S4)"""
from __future__ import annotations

import pathlib

import openpyxl
import pytest
import yaml
from click.testing import CliRunner

from intelligent_regression_optimizer.cli import main

TEMPLATES = pathlib.Path(__file__).parent.parent / "templates"
REPO_TMP = pathlib.Path(__file__).parent / ".tmp"


def _save_tmp(wb: openpyxl.Workbook, name: str) -> pathlib.Path:
    REPO_TMP.mkdir(parents=True, exist_ok=True)
    p = REPO_TMP / name
    wb.save(str(p))
    return p


def _make_minimal_wb(rows: list[list] | None = None) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tests"
    ws.append(["ID", "Name", "Layer", "Coverage Areas",
               "Execution Time (secs)", "Flakiness Rate",
               "Failure Count (30d)", "Automated", "Tags"])
    for row in (rows or [["T-01", "test one", "unit", "ServiceA", 10, 0.01, 0, "true", "smoke"]]):
        ws.append(row)
    return wb


# ---------------------------------------------------------------------------
# Happy path — stdout
# ---------------------------------------------------------------------------

class TestImportTestsStdout:
    def test_valid_xlsx_exits_0(self):
        runner = CliRunner()
        result = runner.invoke(main, ["import-tests", str(TEMPLATES / "test_suite_template.xlsx")])
        assert result.exit_code == 0

    def test_output_contains_test_suite_key(self):
        runner = CliRunner()
        result = runner.invoke(main, ["import-tests", str(TEMPLATES / "test_suite_template.xlsx")])
        assert "test_suite:" in result.output

    def test_output_does_not_contain_sprint_context_stub(self):
        runner = CliRunner()
        result = runner.invoke(main, ["import-tests", str(TEMPLATES / "test_suite_template.xlsx")])
        assert "sprint_context:" not in result.output

    def test_output_does_not_contain_constraints_stub(self):
        runner = CliRunner()
        result = runner.invoke(main, ["import-tests", str(TEMPLATES / "test_suite_template.xlsx")])
        assert "constraints:" not in result.output

    def test_output_is_valid_yaml(self):
        runner = CliRunner()
        result = runner.invoke(main, ["import-tests", str(TEMPLATES / "test_suite_template.xlsx")])
        parsed = yaml.safe_load(result.output)
        assert isinstance(parsed, dict)
        assert "test_suite" in parsed

    def test_all_template_tests_in_output(self):
        runner = CliRunner()
        result = runner.invoke(main, ["import-tests", str(TEMPLATES / "test_suite_template.xlsx")])
        parsed = yaml.safe_load(result.output)
        assert len(parsed["test_suite"]) == 5


# ---------------------------------------------------------------------------
# --output flag
# ---------------------------------------------------------------------------

class TestImportTestsOutputFlag:
    def test_output_flag_writes_file(self):
        out = REPO_TMP / "import_out.yaml"
        REPO_TMP.mkdir(parents=True, exist_ok=True)
        runner = CliRunner()
        result = runner.invoke(
            main,
            ["import-tests", str(TEMPLATES / "test_suite_template.xlsx"), "--output", str(out)],
        )
        assert result.exit_code == 0
        assert out.exists()
        parsed = yaml.safe_load(out.read_text(encoding="utf-8"))
        assert "test_suite" in parsed

    def test_output_flag_nothing_printed_to_stdout(self):
        out = REPO_TMP / "import_quiet.yaml"
        REPO_TMP.mkdir(parents=True, exist_ok=True)
        runner = CliRunner()
        result = runner.invoke(
            main,
            ["import-tests", str(TEMPLATES / "test_suite_template.xlsx"), "--output", str(out)],
        )
        # stdout should be empty or just a confirmation, not the full YAML
        assert "test_suite:" not in result.output


# ---------------------------------------------------------------------------
# --sheet flag
# ---------------------------------------------------------------------------

class TestImportTestsSheetFlag:
    def test_sheet_flag_selects_named_sheet(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Q2 Tests"
        ws.append(["ID", "Name", "Layer", "Coverage Areas",
                   "Execution Time (secs)", "Flakiness Rate",
                   "Failure Count (30d)", "Automated", "Tags"])
        ws.append(["T-Q2", "q2 test", "integration", "ServiceB", 30, 0.02, 0, "true", ""])
        p = _save_tmp(wb, "sheet_flag.xlsx")
        runner = CliRunner()
        result = runner.invoke(main, ["import-tests", str(p), "--sheet", "Q2 Tests"])
        assert result.exit_code == 0
        parsed = yaml.safe_load(result.output)
        assert parsed["test_suite"][0]["id"] == "T-Q2"


# ---------------------------------------------------------------------------
# Error paths
# ---------------------------------------------------------------------------

class TestImportTestsErrors:
    def test_missing_file_exits_2(self):
        runner = CliRunner()
        result = runner.invoke(main, ["import-tests", "nonexistent.xlsx"])
        assert result.exit_code == 2

    def test_bad_extension_exits_2(self):
        csv_path = REPO_TMP / "tests.csv"
        REPO_TMP.mkdir(parents=True, exist_ok=True)
        csv_path.write_text("id,name\n")
        runner = CliRunner()
        result = runner.invoke(main, ["import-tests", str(csv_path)])
        assert result.exit_code == 2

    def test_missing_required_column_exits_2(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tests"
        ws.append(["Name", "Layer"])           # no ID column
        ws.append(["test one", "unit"])
        p = _save_tmp(wb, "no_id_col.xlsx")
        runner = CliRunner()
        result = runner.invoke(main, ["import-tests", str(p)])
        assert result.exit_code == 2

    def test_invalid_cell_value_exits_2(self):
        wb = _make_minimal_wb(rows=[["T-01", "test", "unit", "ServiceA", 10, 9.99, 0, "true", ""]])
        p = _save_tmp(wb, "bad_flakiness.xlsx")
        runner = CliRunner()
        result = runner.invoke(main, ["import-tests", str(p)])
        assert result.exit_code == 2
