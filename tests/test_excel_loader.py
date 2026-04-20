"""Tests for excel_loader.py — written RED before implementation. (S2)"""
from __future__ import annotations

import pathlib

import openpyxl
import pytest

from intelligent_regression_optimizer.excel_loader import (
    ExcelLoaderError,
    load_excel,
)

TEMPLATES = pathlib.Path(__file__).parent.parent / "templates"
REPO_TMP = pathlib.Path(__file__).parent / ".tmp"


def _make_wb(rows: list[list], *, sheet_name: str = "Tests", headers: list[str] | None = None) -> openpyxl.Workbook:
    """Create an in-memory workbook with given rows under sheet_name."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    default_headers = [
        "ID", "Name", "Layer", "Coverage Areas",
        "Execution Time (secs)", "Flakiness Rate",
        "Failure Count (30d)", "Automated", "Tags",
    ]
    ws.append(headers if headers is not None else default_headers)
    for row in rows:
        ws.append(row)
    return wb


def _save_tmp(wb: openpyxl.Workbook, name: str) -> pathlib.Path:
    REPO_TMP.mkdir(parents=True, exist_ok=True)
    p = REPO_TMP / name
    wb.save(str(p))
    return p


# ---------------------------------------------------------------------------
# Happy path — template
# ---------------------------------------------------------------------------

class TestTemplateLoads:
    def test_template_loads_successfully(self):
        tests = load_excel(str(TEMPLATES / "test_suite_template.xlsx"))
        assert isinstance(tests, list)
        assert len(tests) == 5

    def test_template_ids_present(self):
        tests = load_excel(str(TEMPLATES / "test_suite_template.xlsx"))
        ids = [t["id"] for t in tests]
        assert "TEST-001" in ids
        assert "TEST-005" in ids

    def test_template_all_required_fields_present(self):
        tests = load_excel(str(TEMPLATES / "test_suite_template.xlsx"))
        required = {"id", "name", "layer", "coverage_areas", "execution_time_secs", "flakiness_rate"}
        for t in tests:
            assert required.issubset(t.keys()), f"Missing fields in {t}"

    def test_coverage_areas_parsed_as_list(self):
        tests = load_excel(str(TEMPLATES / "test_suite_template.xlsx"))
        for t in tests:
            assert isinstance(t["coverage_areas"], list)

    def test_tags_parsed_as_list(self):
        tests = load_excel(str(TEMPLATES / "test_suite_template.xlsx"))
        for t in tests:
            assert isinstance(t["tags"], list)

    def test_automated_parsed_as_bool(self):
        tests = load_excel(str(TEMPLATES / "test_suite_template.xlsx"))
        for t in tests:
            assert isinstance(t["automated"], bool)

    def test_flakiness_rate_parsed_as_float(self):
        tests = load_excel(str(TEMPLATES / "test_suite_template.xlsx"))
        for t in tests:
            assert isinstance(t["flakiness_rate"], float)

    def test_execution_time_parsed_as_number(self):
        tests = load_excel(str(TEMPLATES / "test_suite_template.xlsx"))
        for t in tests:
            assert isinstance(t["execution_time_secs"], (int, float))


# ---------------------------------------------------------------------------
# Column order flexibility
# ---------------------------------------------------------------------------

class TestColumnOrderFlexibility:
    def test_columns_in_different_order(self):
        wb = _make_wb(
            headers=["Name", "ID", "Tags", "Layer", "Coverage Areas",
                     "Flakiness Rate", "Execution Time (secs)", "Failure Count (30d)", "Automated"],
            rows=[["payment test", "T-01", "", "e2e", "ServiceA", 0.05, 60, 0, "true"]],
        )
        p = _save_tmp(wb, "col_order.xlsx")
        tests = load_excel(str(p))
        assert tests[0]["id"] == "T-01"
        assert tests[0]["name"] == "payment test"


# ---------------------------------------------------------------------------
# Multi-sheet handling
# ---------------------------------------------------------------------------

class TestMultiSheet:
    def test_auto_selects_tests_sheet(self):
        wb = openpyxl.Workbook()
        # First sheet (not Tests)
        ws1 = wb.active
        ws1.title = "Summary"
        ws1.append(["some", "other", "data"])
        # Second sheet named Tests
        ws2 = wb.create_sheet("Tests")
        ws2.append(["ID", "Name", "Layer", "Coverage Areas",
                    "Execution Time (secs)", "Flakiness Rate",
                    "Failure Count (30d)", "Automated", "Tags"])
        ws2.append(["T-01", "test one", "unit", "ServiceA", 10, 0.01, 0, "true", ""])
        p = _save_tmp(wb, "multi_sheet.xlsx")
        tests = load_excel(str(p))
        assert len(tests) == 1
        assert tests[0]["id"] == "T-01"

    def test_explicit_sheet_name(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Q2 Tests"
        ws.append(["ID", "Name", "Layer", "Coverage Areas",
                   "Execution Time (secs)", "Flakiness Rate",
                   "Failure Count (30d)", "Automated", "Tags"])
        ws.append(["T-02", "test two", "integration", "ServiceB", 30, 0.0, 0, "true", ""])
        p = _save_tmp(wb, "named_sheet.xlsx")
        tests = load_excel(str(p), sheet="Q2 Tests")
        assert tests[0]["id"] == "T-02"

    def test_explicit_sheet_not_found_raises(self):
        wb = _make_wb(rows=[["T-01", "t", "unit", "A", 10, 0.0, 0, "true", ""]])
        p = _save_tmp(wb, "sheet_notfound.xlsx")
        with pytest.raises(ExcelLoaderError, match="Sheet.*not found"):
            load_excel(str(p), sheet="NonExistentSheet")

    def test_falls_back_to_first_sheet_when_no_tests_sheet(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MyTests"
        ws.append(["ID", "Name", "Layer", "Coverage Areas",
                   "Execution Time (secs)", "Flakiness Rate",
                   "Failure Count (30d)", "Automated", "Tags"])
        ws.append(["T-03", "third", "e2e", "ServiceC", 60, 0.02, 1, "true", ""])
        p = _save_tmp(wb, "first_sheet_fallback.xlsx")
        tests = load_excel(str(p))
        assert tests[0]["id"] == "T-03"


# ---------------------------------------------------------------------------
# Fuzzy column matching
# ---------------------------------------------------------------------------

class TestFuzzyColumnMatching:
    def test_lowercase_headers_accepted(self):
        wb = _make_wb(
            headers=["id", "name", "layer", "coverage areas",
                     "execution time (secs)", "flakiness rate",
                     "failure count (30d)", "automated", "tags"],
            rows=[["T-01", "test", "unit", "ServiceA", 10, 0.0, 0, "true", ""]],
        )
        p = _save_tmp(wb, "lower_headers.xlsx")
        tests = load_excel(str(p))
        assert tests[0]["id"] == "T-01"

    def test_snake_case_headers_accepted(self):
        wb = _make_wb(
            headers=["id", "name", "layer", "coverage_areas",
                     "execution_time_secs", "flakiness_rate",
                     "failure_count_last_30d", "automated", "tags"],
            rows=[["T-01", "test", "unit", "ServiceA", 10, 0.0, 0, "true", ""]],
        )
        p = _save_tmp(wb, "snake_headers.xlsx")
        tests = load_excel(str(p))
        assert tests[0]["id"] == "T-01"


# ---------------------------------------------------------------------------
# Header row detection
# ---------------------------------------------------------------------------

class TestHeaderRowDetection:
    def test_header_in_row_2(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tests"
        ws.append(["This file was exported from TestRail", None, None, None, None, None, None, None, None])
        ws.append(["ID", "Name", "Layer", "Coverage Areas",
                   "Execution Time (secs)", "Flakiness Rate",
                   "Failure Count (30d)", "Automated", "Tags"])
        ws.append(["T-01", "test one", "unit", "ServiceA", 10, 0.01, 0, "true", ""])
        p = _save_tmp(wb, "header_row2.xlsx")
        tests = load_excel(str(p))
        assert len(tests) == 1
        assert tests[0]["id"] == "T-01"


# ---------------------------------------------------------------------------
# Validation — missing required columns
# ---------------------------------------------------------------------------

class TestMissingRequiredColumns:
    def test_missing_id_column_raises(self):
        wb = _make_wb(
            headers=["Name", "Layer", "Coverage Areas",
                     "Execution Time (secs)", "Flakiness Rate",
                     "Failure Count (30d)", "Automated", "Tags"],
            rows=[["test", "unit", "ServiceA", 10, 0.0, 0, "true", ""]],
        )
        p = _save_tmp(wb, "no_id.xlsx")
        with pytest.raises(ExcelLoaderError, match="[Ii][Dd]"):
            load_excel(str(p))

    def test_missing_layer_column_raises(self):
        wb = _make_wb(
            headers=["ID", "Name", "Coverage Areas",
                     "Execution Time (secs)", "Flakiness Rate",
                     "Failure Count (30d)", "Automated", "Tags"],
            rows=[["T-01", "test", "ServiceA", 10, 0.0, 0, "true", ""]],
        )
        p = _save_tmp(wb, "no_layer.xlsx")
        with pytest.raises(ExcelLoaderError, match="[Ll]ayer"):
            load_excel(str(p))


# ---------------------------------------------------------------------------
# Validation — invalid cell values
# ---------------------------------------------------------------------------

class TestCellValidation:
    def test_flakiness_above_1_raises(self):
        wb = _make_wb(rows=[["T-01", "test", "unit", "ServiceA", 10, 1.5, 0, "true", ""]])
        p = _save_tmp(wb, "flakiness_invalid.xlsx")
        with pytest.raises(ExcelLoaderError, match="[Rr]ow 2|flakiness"):
            load_excel(str(p))

    def test_negative_execution_time_raises(self):
        wb = _make_wb(rows=[["T-01", "test", "unit", "ServiceA", -5, 0.0, 0, "true", ""]])
        p = _save_tmp(wb, "exec_time_invalid.xlsx")
        with pytest.raises(ExcelLoaderError, match="[Rr]ow 2|execution"):
            load_excel(str(p))

    def test_invalid_layer_raises(self):
        wb = _make_wb(rows=[["T-01", "test", "invalid_layer", "ServiceA", 10, 0.0, 0, "true", ""]])
        p = _save_tmp(wb, "layer_invalid.xlsx")
        with pytest.raises(ExcelLoaderError, match="[Rr]ow 2|layer|invalid_layer"):
            load_excel(str(p))

    def test_empty_id_raises(self):
        wb = _make_wb(rows=[["", "test", "unit", "ServiceA", 10, 0.0, 0, "true", ""]])
        p = _save_tmp(wb, "empty_id.xlsx")
        with pytest.raises(ExcelLoaderError, match="[Rr]ow 2|[Ii][Dd]"):
            load_excel(str(p))


# ---------------------------------------------------------------------------
# Edge cases — empty and merged cells
# ---------------------------------------------------------------------------

class TestEdgeCases:
    def test_empty_sheet_headers_only_raises(self):
        wb = _make_wb(rows=[])
        p = _save_tmp(wb, "empty.xlsx")
        with pytest.raises(ExcelLoaderError, match="no data rows"):
            load_excel(str(p))

    def test_trailing_empty_rows_ignored(self):
        wb = _make_wb(rows=[
            ["T-01", "test one", "unit", "ServiceA", 10, 0.01, 0, "true", ""],
            [None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None],
        ])
        p = _save_tmp(wb, "trailing_empty.xlsx")
        tests = load_excel(str(p))
        assert len(tests) == 1

    def test_whitespace_stripped_from_values(self):
        wb = _make_wb(rows=[["  T-01  ", "  test name  ", "unit", "ServiceA", 10, 0.0, 0, "true", ""]])
        p = _save_tmp(wb, "whitespace.xlsx")
        tests = load_excel(str(p))
        assert tests[0]["id"] == "T-01"
        assert tests[0]["name"] == "test name"

    def test_file_not_found_raises(self):
        with pytest.raises(ExcelLoaderError, match="not found|File"):
            load_excel("nonexistent.xlsx")

    def test_non_xlsx_extension_raises(self):
        p = REPO_TMP / "test.csv"
        REPO_TMP.mkdir(parents=True, exist_ok=True)
        p.write_text("id,name\n")
        with pytest.raises(ExcelLoaderError, match=".csv|format"):
            load_excel(str(p))

    def test_merged_cells_raises(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tests"
        ws.append(["ID", "Name", "Layer", "Coverage Areas",
                   "Execution Time (secs)", "Flakiness Rate",
                   "Failure Count (30d)", "Automated", "Tags"])
        ws.append(["T-01", "test one", "unit", "ServiceA", 10, 0.01, 0, "true", ""])
        ws.merge_cells("B2:C2")
        p = _save_tmp(wb, "merged.xlsx")
        with pytest.raises(ExcelLoaderError, match="[Mm]erged"):
            load_excel(str(p))

    def test_no_recognisable_header_row_raises(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tests"
        # 6 rows of junk — no recognisable headers
        for _ in range(6):
            ws.append(["foo", "bar", "baz"])
        p = _save_tmp(wb, "no_header.xlsx")
        with pytest.raises(ExcelLoaderError, match="[Hh]eader"):
            load_excel(str(p))


# ---------------------------------------------------------------------------
# Coverage gap tests — _parse_bool branches and non-string cell types
# ---------------------------------------------------------------------------

class TestParseBoolBranches:
    def test_integer_1_automated_true(self):
        wb = _make_wb(rows=[["T-01", "test", "unit", "ServiceA", 10, 0.0, 0, 1, ""]])
        p = _save_tmp(wb, "auto_int1.xlsx")
        tests = load_excel(str(p))
        assert tests[0]["automated"] is True

    def test_integer_0_automated_false(self):
        wb = _make_wb(rows=[["T-01", "test", "unit", "ServiceA", 10, 0.0, 0, 0, ""]])
        p = _save_tmp(wb, "auto_int0.xlsx")
        tests = load_excel(str(p))
        assert tests[0]["automated"] is False

    def test_unrecognised_automated_value_raises(self):
        """Malformed Automated value 'maybe' must raise, not silently default to True."""
        wb = _make_wb(rows=[["T-01", "test", "unit", "ServiceA", 10, 0.0, 0, "maybe", ""]])
        p = _save_tmp(wb, "auto_maybe.xlsx")
        with pytest.raises(ExcelLoaderError, match="[Rr]ow 2|[Aa]utomated"):
            load_excel(str(p))

    def test_non_numeric_execution_time_raises(self):
        wb = _make_wb(rows=[["T-01", "test", "unit", "ServiceA", "fast", 0.0, 0, "true", ""]])
        p = _save_tmp(wb, "exec_nonnumeric.xlsx")
        with pytest.raises(ExcelLoaderError, match="[Rr]ow 2|[Ee]xecution"):
            load_excel(str(p))

    def test_non_numeric_flakiness_raises(self):
        wb = _make_wb(rows=[["T-01", "test", "unit", "ServiceA", 10, "high", 0, "true", ""]])
        p = _save_tmp(wb, "flakiness_nonnumeric.xlsx")
        with pytest.raises(ExcelLoaderError, match="[Rr]ow 2|[Ff]lakiness"):
            load_excel(str(p))

    def test_empty_name_raises(self):
        wb = _make_wb(rows=[["T-01", "", "unit", "ServiceA", 10, 0.0, 0, "true", ""]])
        p = _save_tmp(wb, "empty_name.xlsx")
        with pytest.raises(ExcelLoaderError, match="[Rr]ow 2|[Nn]ame"):
            load_excel(str(p))

    def test_blank_coverage_areas_raises(self):
        """Blank required Coverage Areas must raise, not silently produce empty list."""
        wb = _make_wb(rows=[["T-01", "test", "unit", "", 10, 0.0, 0, "true", ""]])
        p = _save_tmp(wb, "blank_coverage.xlsx")
        with pytest.raises(ExcelLoaderError, match="[Rr]ow 2|[Cc]overage"):
            load_excel(str(p))

    def test_blank_execution_time_raises(self):
        """Blank required Execution Time must raise, not silently produce 0.0."""
        wb = _make_wb(rows=[["T-01", "test", "unit", "ServiceA", None, 0.0, 0, "true", ""]])
        p = _save_tmp(wb, "blank_exec_time.xlsx")
        with pytest.raises(ExcelLoaderError, match="[Rr]ow 2|[Ee]xecution"):
            load_excel(str(p))

    def test_blank_flakiness_rate_raises(self):
        """Blank required Flakiness Rate must raise, not silently produce 0.0."""
        wb = _make_wb(rows=[["T-01", "test", "unit", "ServiceA", 10, None, 0, "true", ""]])
        p = _save_tmp(wb, "blank_flakiness.xlsx")
        with pytest.raises(ExcelLoaderError, match="[Rr]ow 2|[Ff]lakiness"):
            load_excel(str(p))

    def test_malformed_failure_count_raises(self):
        """Non-integer Failure Count must raise, not silently produce 0."""
        wb = _make_wb(rows=[["T-01", "test", "unit", "ServiceA", 10, 0.0, "many", "true", ""]])
        p = _save_tmp(wb, "bad_failure_count.xlsx")
        with pytest.raises(ExcelLoaderError, match="[Rr]ow 2|[Ff]ailure"):
            load_excel(str(p))


# ---------------------------------------------------------------------------
# Passthrough columns — Priority, External ID, Owner, Module
# ---------------------------------------------------------------------------

class TestPassthroughColumns:
    def test_template_has_passthrough_fields(self):
        tests = load_excel(str(TEMPLATES / "test_suite_template.xlsx"))
        t = tests[0]
        assert t["priority"] == "P0"
        assert t["external_id"] == "JIRA-101"
        assert t["owner"] == "alice"
        assert t["module"] == "Payment"

    def test_passthrough_fields_present_when_absent_in_sheet(self):
        """Columns not in sheet → fields returned as None."""
        wb = _make_wb(rows=[["T-01", "test", "unit", "ServiceA", 10, 0.0, 0, "true", ""]])
        p = _save_tmp(wb, "no_passthrough.xlsx")
        tests = load_excel(str(p))
        assert tests[0]["priority"] is None
        assert tests[0]["external_id"] is None
        assert tests[0]["owner"] is None
        assert tests[0]["module"] is None

    def test_passthrough_fields_empty_cell_returns_none(self):
        """Empty cells in passthrough columns → None."""
        headers = [
            "ID", "Name", "Layer", "Coverage Areas",
            "Execution Time (secs)", "Flakiness Rate",
            "Failure Count (30d)", "Automated", "Tags",
            "Priority", "External ID", "Owner", "Module",
        ]
        wb = _make_wb(
            headers=headers,
            rows=[["T-01", "test", "unit", "ServiceA", 10, 0.0, 0, "true", "", "", "", "", ""]],
        )
        p = _save_tmp(wb, "passthrough_empty.xlsx")
        tests = load_excel(str(p))
        assert tests[0]["priority"] is None
        assert tests[0]["external_id"] is None
        assert tests[0]["owner"] is None
        assert tests[0]["module"] is None

    def test_passthrough_values_preserved_as_strings(self):
        headers = [
            "ID", "Name", "Layer", "Coverage Areas",
            "Execution Time (secs)", "Flakiness Rate",
            "Failure Count (30d)", "Automated", "Tags",
            "Priority", "External ID", "Owner", "Module",
        ]
        wb = _make_wb(
            headers=headers,
            rows=[["T-01", "test", "unit", "ServiceA", 10, 0.0, 0, "true", "",
                   "P1", "JIRA-999", "qeteam", "OrderManagement"]],
        )
        p = _save_tmp(wb, "passthrough_values.xlsx")
        tests = load_excel(str(p))
        assert tests[0]["priority"] == "P1"
        assert tests[0]["external_id"] == "JIRA-999"
        assert tests[0]["owner"] == "qeteam"
        assert tests[0]["module"] == "OrderManagement"

    def test_passthrough_whitespace_stripped(self):
        headers = [
            "ID", "Name", "Layer", "Coverage Areas",
            "Execution Time (secs)", "Flakiness Rate",
            "Failure Count (30d)", "Automated", "Tags",
            "Priority", "External ID", "Owner", "Module",
        ]
        wb = _make_wb(
            headers=headers,
            rows=[["T-01", "test", "unit", "ServiceA", 10, 0.0, 0, "true", "",
                   "  P2  ", "  TR-42  ", "  alice  ", "  Payment  "]],
        )
        p = _save_tmp(wb, "passthrough_whitespace.xlsx")
        tests = load_excel(str(p))
        assert tests[0]["priority"] == "P2"
        assert tests[0]["external_id"] == "TR-42"
        assert tests[0]["owner"] == "alice"
        assert tests[0]["module"] == "Payment"

    def test_module_fuzzy_alias_component(self):
        headers = [
            "ID", "Name", "Layer", "Coverage Areas",
            "Execution Time (secs)", "Flakiness Rate",
            "Failure Count (30d)", "Automated", "Tags",
            "Component",
        ]
        wb = _make_wb(
            headers=headers,
            rows=[["T-01", "test", "unit", "ServiceA", 10, 0.0, 0, "true", "", "PaymentGateway"]],
        )
        p = _save_tmp(wb, "module_alias_component.xlsx")
        tests = load_excel(str(p))
        assert tests[0]["module"] == "PaymentGateway"

    def test_external_id_fuzzy_alias_jira(self):
        headers = [
            "ID", "Name", "Layer", "Coverage Areas",
            "Execution Time (secs)", "Flakiness Rate",
            "Failure Count (30d)", "Automated", "Tags",
            "Jira",
        ]
        wb = _make_wb(
            headers=headers,
            rows=[["T-01", "test", "unit", "ServiceA", 10, 0.0, 0, "true", "", "PAY-55"]],
        )
        p = _save_tmp(wb, "extid_alias_jira.xlsx")
        tests = load_excel(str(p))
        assert tests[0]["external_id"] == "PAY-55"
